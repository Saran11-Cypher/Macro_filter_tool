import os
import pandas as pd
from openpyxl import load_workbook

# Define paths
EXCEL_FILE = "E:\\PYTHON\\Django\\Workspace\\Macro_Generator\\env\\Macro_Functional_Excel.xlsx"  # Update with your actual file path
UPLOAD_FOLDER = "E:\\PYTHON\\ServiceCategory"  # Change to the folder containing uploaded files

# Ensure the folder exists
if not os.path.exists(UPLOAD_FOLDER):
    print(f"❌ Error: Folder '{UPLOAD_FOLDER}' does not exist.")
    exit()

# Load workbook and sheets
wb = load_workbook(EXCEL_FILE)
ws_main = wb["Main"]
ws_bal = wb["Business Approved List"]

# List all uploaded files
uploaded_files = [f for f in os.listdir(UPLOAD_FOLDER) if os.path.isfile(os.path.join(UPLOAD_FOLDER, f))]
file_count = len(uploaded_files)

# Update "Main" sheet with file count
ws_main.append(["Service_Category", file_count, "Pending", "Pending", "Pending"])

# Load "Business Approved List" into a DataFrame
df_bal = pd.read_excel(EXCEL_FILE, sheet_name="Business Approved List")

# Define the correct config load order
config_load_order = [
    "ValueList", "AttributeType", "UserDefinedTerm", "LineOfBusiness",
    "Product", "ServiceCategory", "BenefitNetwork", "NetworkDefinitionComponent",
    "BenefitPlanComponent", "WrapAroundBenefitPlan", "BenefitPlanRider",
    "BenefitPlanTemplate", "Account", "BenefitPlan", "AccountPlanSelection"
]

# Convert "Config Type" to string to avoid errors
df_bal["Config Type"] = df_bal["Config Type"].astype(str)

# Assign order index based on predefined sequence
df_bal["Order"] = df_bal["Config Type"].apply(lambda x: config_load_order.index(x) if x in config_load_order else -1)

# Validate order: If not in increasing sequence, show error and exit
if not df_bal["Order"].is_monotonic_increasing:
    print("❌ Error: Invalid Order! Please arrange the data correctly.")
    exit()

# Remove the temporary "Order" column (not needed in final output)
df_bal.drop(columns=["Order"], inplace=True)

# Check for HRL availability and update file paths in the same row
for index, row in df_bal.iterrows():
    config_name = str(row["Config Name"])
    matching_files = [f for f in uploaded_files if config_name in f]

    if matching_files:
        df_bal.at[index, "HRL Available?"] = "HRL Found"
        df_bal.at[index, "File Name is correct in export sheet"] = os.path.join(UPLOAD_FOLDER, matching_files[0])
    else:
        df_bal.at[index, "HRL Available?"] = "Not Found"

# Save updates back to Excel without adding extra rows
for row_idx, row in df_bal.iterrows():
    for col_idx, value in enumerate(row):
        ws_bal.cell(row=row_idx+2, column=col_idx+1, value=value)  # Update same row

# Save the final workbook
wb.save(EXCEL_FILE)

print("✅ Excel file updated successfully!")
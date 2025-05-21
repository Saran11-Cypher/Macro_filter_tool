import os
import pandas as pd
import re
from datetime import datetime
import shutil
from openpyxl import load_workbook

EXCEL_FILE = "C:\\Users\\n925072\\Downloads\\MacroFile_Conversion-master\\MacroFile_Conversion-master\\New folder\\convertor\\Macro_Functional_Excel.xlsx"  # Update with your actual file path
UPLOAD_FOLDER = "C:\\1"  # Change to the folder containing uploaded files
DATE_STAMP = datetime.now().strftime("%Y%m%d_%H%M%S")  # New parent folder with date and time stamp
HRL_PARENT_FOLDER = f"C:\\Datas\\HRLS_{DATE_STAMP}"
# Ensure the folder exists
if not os.path.exists(UPLOAD_FOLDER):
    print(f"❌ Error: Folder '{UPLOAD_FOLDER}' does not exist.")
    exit()

# Load workbook and sheets
wb = load_workbook(EXCEL_FILE)
ws_main = wb["Main"]
ws_bal = wb["Business Approved List"]

# Define the correct config load order
config_load_order = [
    "ValueList", "AttributeType", "UserDefinedTerm", "LineOfBusiness",
    "Product", "ServiceCategory", "BenefitNetwork", "NetworkDefinitionComponent",
    "BenefitPlanComponent", "WrapAroundBenefitPlan", "BenefitPlanRider",
    "BenefitPlanTemplate", "Account", "BenefitPlan", "AccountPlanSelection"
]

def trim_suffix(filename):
    return re.sub(r'\.\d{4}-\d{2}-\d{2}\..*$', '', filename)

# Function to normalize and clean text
def normalize_text(text):
    """Removes special characters, converts to lowercase, and standardizes spaces/hyphens."""
    return re.sub(r'[^a-zA-Z0-9.]','', str(text)).strip().lower()

# Load "Business Approved List" into a DataFrame
df_bal = pd.read_excel(EXCEL_FILE, sheet_name="Business Approved List", dtype=str)  # Ensure all columns are strings
df_bal["Config Type"] = df_bal["Config Type"].astype(str).apply(normalize_text)

# Get the config types mentioned in "Business Approved List"
approved_config_types = set(df_bal["Config Type"].dropna().unique())

# Get list of subfolders inside the parent folder
available_folders = {normalize_text(f): os.path.join(UPLOAD_FOLDER, f) 
                     for f in os.listdir(UPLOAD_FOLDER) if os.path.isdir(os.path.join(UPLOAD_FOLDER, f))}

# Select only folders present in both the config load order and "Business Approved List"
selected_folders = {config: path for config, path in available_folders.items() if config in approved_config_types}

if not selected_folders:
    print("❌ Error: No matching config folders found inside the parent folder.")
    exit()

# Process each selected folder
for config_type, folder_path in selected_folders.items():
    uploaded_files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
    file_count = len(uploaded_files)

    # Update the "Main" sheet dynamically
    ws_main.append([config_type, file_count, "Pending", "Pending", "Pending"])

# Assign order dynamically based on available configurations only
df_bal["Order"] = df_bal["Config Type"].apply(lambda x: config_load_order.index(x) if x in config_load_order else -1)

# Validate order: If not in increasing sequence, show error and exit
valid_orders = df_bal[df_bal["Order"] >= 0]["Order"]

if not valid_orders.is_monotonic_increasing:
    print("❌ Error: Invalid Order! Please arrange the data correctly.")
    exit()

# Remove the temporary "Order" column (not needed in final output)
df_bal.drop(columns=["Order"], inplace=True)

# Function to match config names with uploaded files
def find_matching_file(config_type, config_name, folder_path):
    if "&" in config_name:
        config_name = config_name.replace("&", "and")
    """Find an exact match for the given Config Type and Config Name in the folder."""
    normalized_config_type = normalize_text(config_type)
    normalized_config_name = normalize_text(config_name)

    # Store all filenames in a normalized format without date suffix
    normalized_files = {
        normalize_text(trim_suffix(f)): f 
        for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))
    }

    # Debug Output: Print only normalized filenames (keys)
    # print("\n--- Normalized Filenames Without Date (Keys Only) ---")
    # for norm_file in normalized_files.keys():
    #     print(norm_file)

    # Construct the expected pattern (without date & extension)
    expected_pattern = f"{normalized_config_type}.{normalized_config_name}"

    # Debug Output: Print the expected pattern
    print(f"\nExpected Pattern: {expected_pattern}")

    # Check for an exact match
    for norm_file, original_file in normalized_files.items():
        if norm_file == expected_pattern:  # Ensure exact match
            print(f"\n✅ Match Found: {original_file}")
            return original_file  # Return the actual filename

    print("\n❌ No exact match found")
    return None  # No exact match found
# Check for HRL availability and update DataFrame
for index, row in df_bal.iterrows():
    config_type = row["Config Type"]
    config_name = row["Config Name"]

    if pd.isna(config_name) or not str(config_name).strip():
        continue

    if config_type in selected_folders:
        matching_file = find_matching_file(config_type,config_name, selected_folders[config_type])

        if matching_file:
            df_bal.at[index, "HRL Available?"] = "HRL Found"
            source_path = os.path.join(selected_folders[config_type], matching_file)
            target_folder = os.path.join(HRL_PARENT_FOLDER, config_type)
            os.makedirs(target_folder, exist_ok=True)
            target_path = os.path.join(target_folder, matching_file)
            
            shutil.copy2(source_path, target_path)  # Copy HRL file to the new parent folder
            df_bal.at[index, "File Name is correct in export sheet"] = source_path  # Keep original path
        else:
            df_bal.at[index, "HRL Available?"] = "Not Found"

for row_idx, row in df_bal.iterrows():
    for col_idx, value in enumerate(row):
        ws_bal.cell(row=row_idx+2, column=col_idx+1, value=str(value))

wb.save(EXCEL_FILE)
print(f"✅ HRL files copied to '{HRL_PARENT_FOLDER}' and Excel file updated successfully!")
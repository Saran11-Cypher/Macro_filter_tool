import os, re, traceback, shutil, time
import pandas as pd
from collections import defaultdict
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from django.conf import settings
import os, re, traceback, shutil
from django.core.cache import cache
def is_file_locked(filepath):
    if not os.path.exists(filepath):
        return False
    try:
        with open(filepath, 'a'):
            return False
    except OSError:
        return True
    
# Helper functions
def normalize_text(text):
    return re.sub(r'[^a-zA-Z0-9]', '', str(text)).strip().lower()

def extract_date_from_filename(filename):
    match = re.search(r'\.(\d{4}-\d{2}-\d{2})\.', filename)
    if match:
        try:
            return datetime.strptime(match.group(1), "%Y-%m-%d")
        except ValueError:
            return None
    return None

def categorize_files(folder_path):
    single_version_files = {}
    multi_version_files = defaultdict(list)

    for root, dirs, files in os.walk(folder_path):
        for file in files:
            parts = file.split('.')

            # Type 1: Exactly 3 parts — configtype.configname.dateformat.hrl
            if len(parts) == 4:
                config_name = parts[1]

            # Type 2: More than 3 parts — configtype.configname.part1.part2...dateformat.hrl
            elif len(parts) > 4:
                config_name = '.'.join(parts[1:-3])  # Join all between config type and date

            else:
                continue  # Skip invalid files

            normalized_config_name = normalize_text(config_name)
            # print(f"🔍 Processing with normalized config name: {normalized_config_name}")

            if normalized_config_name in single_version_files:
                multi_version_files[normalized_config_name].append(file)
                multi_version_files[normalized_config_name].append(single_version_files.pop(normalized_config_name)[0])
            elif normalized_config_name in multi_version_files:
                multi_version_files[normalized_config_name].append(file)
            else:
                single_version_files[normalized_config_name] = [file]

    return single_version_files, multi_version_files

def find_matching_file(config_name, single_version_files, multi_version_files, selected_version):
    # 🔁 In-place decoding replacements
    replacements = {
        "&": "and",
        "%": "perc",
        "$": "dollar",
    }

    for symbol, replacement in replacements.items():
        config_name = config_name.replace(symbol, replacement)

    normalized_key = normalize_text(config_name)

    if normalized_key in single_version_files:
        return [single_version_files[normalized_key][0]]

    elif normalized_key in multi_version_files:
        candidates = multi_version_files[normalized_key]
        dated = [(f, extract_date_from_filename(f)) for f in candidates]
        dated.sort(key=lambda x: (x[1] or datetime.min))

        if selected_version == 'latest':
            print(f"✅ Latest version selected: {dated[-1][0]}")
            return [dated[-1][0]]
        elif selected_version == 'oldest':
            print(f"✅ Oldest version selected: {dated[0][0]}")
            return [dated[0][0]]
        else:
            print(f"✅ All versions selected: {[f for f, _ in dated]}")
            return [f for f, _ in dated]

    print("❌ No match found.")
    return []


def safe_copy(src, tgt, max_retries=3):
    if not os.path.exists(src):
        print(f"❌ Source file missing: {src}")
        raise FileNotFoundError(f"Source file missing: {src}")

    # ✅ Convert to long path format for Windows
    if os.name == 'nt':
        src = f"\\\\?\\{os.path.abspath(src)}"
        tgt = f"\\\\?\\{os.path.abspath(tgt)}"

    for attempt in range(max_retries):
        try:
            if not os.path.exists(tgt):
                try:
                    os.link(src, tgt)
                except Exception:
                    shutil.copy2(src, tgt)
            return
        except PermissionError as e:
            if attempt < max_retries - 1:
                time.sleep(0.5)
            else:
                raise e



def process_hrl_files(excel_path, upload_folder, version_choice, file_id, progress_callback=None):
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        HRL_PARENT_FOLDER = os.path.join(settings.MEDIA_ROOT, f"HRLS_{timestamp}")
        os.makedirs(HRL_PARENT_FOLDER, exist_ok=True)

        df_bal = pd.read_excel(
            excel_path,
            sheet_name="Business Approved List",
            dtype=str,
            keep_default_na=False,
            na_filter=False
        )
        df_bal.columns = df_bal.columns.str.strip()

        col_map = {col.strip().lower(): col for col in df_bal.columns}
        required_keys = {
            "hrl available?": col_map.get("hrl available?"),
            "file name is correct in export sheet": col_map.get("file name is correct in export sheet")
        }
        if None in required_keys.values():
            raise Exception("Missing required column(s) in Excel sheet")

        hrl_col = required_keys["hrl available?"]
        file_name_col = required_keys["file name is correct in export sheet"]

        df_bal["Config Type"] = df_bal["Config Type"].astype(str).apply(normalize_text)
        approved_config_types = set(df_bal["Config Type"].dropna().unique())

        config_root_path = os.path.join(settings.MEDIA_ROOT, "configs")
        available_folders = {
            normalize_text(f): os.path.join(config_root_path, f)
            for f in os.listdir(config_root_path)
            if os.path.isdir(os.path.join(config_root_path, f))
        }
        selected_folders = {k: v for k, v in available_folders.items() if k in approved_config_types}

        categorized_map = {
            config_type: categorize_files(folder_path)
            for config_type, folder_path in selected_folders.items()
        }

        total_rows = sum(len(df_bal[df_bal["Config Type"] == config_type]) for config_type in selected_folders)
        processed_rows = 0

        for config_type, folder_path in selected_folders.items():
            if cache.get(f"cancel_filtration_{file_id}"):
                print("❌ Filtration cancelled before processing config_type:", config_type)
                if progress_callback:
                    progress_callback(-1)
                return

            single_version_files, multi_version_files = categorized_map[config_type]
            config_type_rows = df_bal[df_bal["Config Type"] == config_type]

            for index, row in config_type_rows.iterrows():
                if cache.get(f"cancel_filtration_{file_id}"):
                    print(f"❌ Cancelled during row index {index} of config_type {config_type}")
                    if progress_callback:
                        progress_callback(-1)
                    return

                config_name = row.get("Config Name")
                if pd.isna(config_name) or not str(config_name).strip():
                    continue

                matches = find_matching_file(config_name, single_version_files, multi_version_files, version_choice)
                matches = list({os.path.basename(m): m for m in matches}.values())

                if matches:
                    df_bal.at[index, hrl_col] = "HRL Found"
                    for match in matches:
                        src = os.path.join(folder_path, match)
                        tgt_folder = os.path.join(HRL_PARENT_FOLDER, config_type)
                        os.makedirs(tgt_folder, exist_ok=True)
                        tgt = os.path.join(tgt_folder, match)
                        safe_copy(src, tgt)

                        existing_val = str(df_bal.at[index, file_name_col]) if pd.notna(df_bal.at[index, file_name_col]) else ""
                        new_val = os.path.relpath(src, upload_folder)
                        df_bal.at[index, file_name_col] = (existing_val + f", {new_val}").strip(", ")
                else:
                    df_bal.at[index, hrl_col] = "Not Found"

                processed_rows += 1
                if progress_callback:
                    progress_callback(int((processed_rows / total_rows) * 95))

        # Save filtered Excel with sheet name
        filtered_excel_path = os.path.join(HRL_PARENT_FOLDER, os.path.basename(excel_path))
        df_bal.to_excel(filtered_excel_path, index=False, sheet_name="Business Approved List")

        # Load workbook and create Main sheet at index 0
        wb = load_workbook(filtered_excel_path)
        ws_main = wb.create_sheet(title="Main", index=0)
        ws_main.append(["Config Type", "Files Found", "Exported", "Imported"])
        for config_type, folder_path in selected_folders.items():
            uploaded_files = [
                f for f in os.listdir(folder_path)
                if os.path.isfile(os.path.join(folder_path, f))
            ]
            ws_main.append([config_type, len(uploaded_files), "Pending", "Pending"])

        # Style the Business Approved List sheet
        ws = wb["Business Approved List"]
        header = [
            str(cell.value).strip().lower()
            for cell in ws[1]
            if cell.value is not None
        ]

        config_type_col = header.index("config type") + 1
        config_name_col = header.index("config name") + 1

        red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        blue = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')

        for row in range(2, ws.max_row + 1):
            cfg_type = normalize_text(ws.cell(row=row, column=config_type_col).value or "")
            cfg_name = normalize_text(ws.cell(row=row, column=config_name_col).value or "")
            if cfg_type in categorized_map:
                single, multi = categorized_map[cfg_type]
                version_count = len(single.get(cfg_name, [])) + len(multi.get(cfg_name, []))
                cell = ws.cell(row=row, column=config_name_col)
                if version_count == 1:
                    cell.fill = red
                elif version_count == 2:
                    cell.fill = green
                elif version_count > 2:
                    cell.fill = blue

        # Force correct sheet order: Main, Business Approved List
        if "Main" in wb.sheetnames and "Business Approved List" in wb.sheetnames:
            main_sheet = wb["Main"]
            bal_sheet = wb["Business Approved List"]
            wb._sheets.remove(main_sheet)
            wb._sheets.remove(bal_sheet)
            wb._sheets = [main_sheet, bal_sheet] + wb._sheets

        wb.save(filtered_excel_path)

        if cache.get(f"cancel_filtration_{file_id}"):
            print("❌ Cancelled before final save.")
            if progress_callback:
                progress_callback(-1)
            return

        if progress_callback:
            progress_callback(100)

        return filtered_excel_path

    except Exception as e:
        print("❌ Exception in process_hrl_files root:", str(e))
        traceback.print_exc()
        if progress_callback:
            progress_callback(-1)
        raise








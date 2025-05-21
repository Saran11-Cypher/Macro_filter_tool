import os
import pandas as pd
import re
from datetime import datetime
import shutil
from openpyxl import load_workbook
from collections import defaultdict
from openpyxl.styles import PatternFill

# Constants
EXCEL_FILE = "C:\\Users\\n925072\\Downloads\\MacroFile_Conversion-master\\MacroFile_Conversion-master\\New folder\\convertor\\Macro_Functional_Excel.xlsx"
UPLOAD_FOLDER = "C:\\1"
DATE_STAMP = datetime.now().strftime("%Y%m%d_%H%M%S")
HRL_PARENT_FOLDER = f"C:\\Datas\\HRLS_{DATE_STAMP}"

# Check upload folder
if not os.path.exists(UPLOAD_FOLDER):
    print(f"❌ Error: Folder '{UPLOAD_FOLDER}' does not exist.")
    exit()

# Load Excel workbook and sheets
wb = load_workbook(EXCEL_FILE)
ws_main = wb["Main"]
ws_bal = wb["Business Approved List"]

# Load config load order
config_load_order = [
    "ValueList", "AttributeType", "UserDefinedTerm", "LineOfBusiness",
    "Product", "ServiceCategory", "BenefitNetwork", "NetworkDefinitionComponent",
    "BenefitPlanComponent", "WrapAroundBenefitPlan", "BenefitPlanRider",
    "BenefitPlanTemplate", "Account", "BenefitPlan", "AccountPlanSelection"
]

# Helper functions
def normalize_text(text):
    return re.sub(r'[^a-zA-Z0-9.]', '', str(text)).strip().lower()

def extract_date_from_filename(filename):
    match = re.search(r'\.(\d{4}-\d{2}-\d{2})\.', filename)
    if match:
        try:
            return datetime.strptime(match.group(1), "%Y-%m-%d")
        except ValueError:
            return None
    return None

def categorize_files(folder_path):
    single_version_files = {}
    multi_version_files = defaultdict(list)

    for root, dirs, files in os.walk(folder_path):
        for file in files:
            parts = file.split('.')
            if len(parts) >= 3:
                config_name = parts[1]
            else:
                continue
            normalized_config_name = normalize_text(config_name)

            if normalized_config_name in single_version_files:
                multi_version_files[normalized_config_name].append(file)
                multi_version_files[normalized_config_name].append(single_version_files.pop(normalized_config_name)[0])
            elif normalized_config_name in multi_version_files:
                multi_version_files[normalized_config_name].append(file)
            else:
                single_version_files[normalized_config_name] = [file]

    return single_version_files, multi_version_files

# Load and normalize approved list
df_bal = pd.read_excel(EXCEL_FILE, sheet_name="Business Approved List", dtype=str)
df_bal["Config Type"] = df_bal["Config Type"].astype(str).apply(normalize_text)
approved_config_types = set(df_bal["Config Type"].dropna().unique())
print(f"✅ Found {len(approved_config_types)} approved config types.")

# Available folders
available_folders = {normalize_text(f): os.path.join(UPLOAD_FOLDER, f)
                     for f in os.listdir(UPLOAD_FOLDER) if os.path.isdir(os.path.join(UPLOAD_FOLDER, f))}
selected_folders = {config: path for config, path in available_folders.items() if config in approved_config_types}

if not selected_folders:
    print("❌ Error: No matching config folders found inside the parent folder.")
    exit()

print(f"✅ Found {len(selected_folders)} matching folders in the upload directory.")
# Process each selected folder
for config_type, folder_path in selected_folders.items():
    uploaded_files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
    file_count = len(uploaded_files)

    # Update the "Main" sheet dynamically
    ws_main.append([config_type, file_count, "Pending", "Pending", "Pending"])
# Global user choice
while True:
    user_choice = input(
        "\n🔎 Do you want to pick the (L)atest, (O)ldest, or (A)ll versions for multi-versions? (L/O/A): "
    ).strip().lower()
    if user_choice in ('l', 'o', 'a'):
        break
    else:
        print("❗ Invalid input. Please type 'L' for latest, 'O' for oldest, or 'A' for all versions.")

selected_version = {'l': 'latest', 'o': 'oldest', 'a': 'all'}[user_choice]
print(f"\n✅ You have selected to pick **{selected_version.upper()}** version(s) for all files.\n")

# Validate config load order
df_bal["Order"] = df_bal["Config Type"].apply(lambda x: config_load_order.index(x) if x in config_load_order else -1)
valid_orders = df_bal[df_bal["Order"] >= 0]["Order"]

if not valid_orders.is_monotonic_increasing:
    print("❌ Error: Invalid Order! Please arrange the data correctly.")
    exit()

df_bal.drop(columns=["Order"], inplace=True)

def find_matching_file(config_name, single_version_files, multi_version_files):
    if "&" in config_name:
        config_name = config_name.replace("&", "and")

    normalized_key = normalize_text(config_name)
    print(f"🔍 Finding matching file for {normalized_key}")

    if normalized_key in single_version_files:
        print(f"✅ Found in single-version files: {single_version_files[normalized_key][0]}")
        return [single_version_files[normalized_key][0]]
    elif normalized_key in multi_version_files:
        candidates = multi_version_files[normalized_key]
        candidates_with_dates = [(file, extract_date_from_filename(file)) for file in candidates]
        candidates_with_dates.sort(key=lambda x: (x[1] or datetime.min))

        if selected_version == 'latest':
            return [candidates_with_dates[-1][0]]
        elif selected_version == 'oldest':
            return [candidates_with_dates[0][0]]
        else:  # all
            return [file for file, _ in candidates_with_dates]

    print(f"❌ No matching file found for {normalized_key}")
    return []

# Main loop
print("🔄 Checking HRL availability and copying files...")

for config_type, folder_path in selected_folders.items():
    print(f"\n📂 Processing Config Type: {config_type}")

    single_version_files, multi_version_files = categorize_files(folder_path)
    config_type_rows = df_bal[df_bal["Config Type"] == config_type]

    for index, row in config_type_rows.iterrows():
        config_name = row["Config Name"]

        if pd.isna(config_name) or not str(config_name).strip():
            continue

        matching_files = find_matching_file(config_name, single_version_files, multi_version_files)

        if matching_files:
            df_bal.at[index, "HRL Available?"] = "HRL Found"
            for matched_file in matching_files:
                source_path = os.path.join(folder_path, matched_file)
                target_folder = os.path.join(HRL_PARENT_FOLDER, config_type)
                os.makedirs(target_folder, exist_ok=True)
                target_path = os.path.join(target_folder, matched_file)
                shutil.copy2(source_path, target_path)

                current_value = df_bal.at[index, "File Name is correct in export sheet"]
                if pd.isna(current_value) or not current_value:
                    df_bal.at[index, "File Name is correct in export sheet"] = source_path
                else:
                    df_bal.at[index, "File Name is correct in export sheet"] += f", {source_path}"
        else:
            df_bal.at[index, "HRL Available?"] = "Not Found"

# Write updates back to Excel
print("🔄 Writing updated DataFrame back to 'Business Approved List' sheet...")
for row_idx, row in df_bal.iterrows():
    for col_idx, value in enumerate(row):
        ws_bal.cell(row=row_idx + 2, column=col_idx + 1, value=str(value))

# Define colors
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')    # Light Red
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')  # Light Green
blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')   # Light Blue

# Find column indexes
header_row = next(ws_bal.iter_rows(min_row=1, max_row=1, values_only=True))
config_type_col = header_row.index("Config Type") + 1
config_name_col = header_row.index("Config Name") + 1

# Apply color coding
for row in range(2, ws_bal.max_row + 1):
    config_type = normalize_text(ws_bal.cell(row=row, column=config_type_col).value)
    config_name = normalize_text(ws_bal.cell(row=row, column=config_name_col).value)

    if config_type in selected_folders:
        folder_path = selected_folders[config_type]
        _, multi_version_files = categorize_files(folder_path)

        # Count total versions for that config name
        total_versions = 0
        if config_name in multi_version_files:
            total_versions = len(multi_version_files[config_name])
        elif config_name:  # It could be single version
            total_versions = 1

        cell = ws_bal.cell(row=row, column=config_name_col)

        if total_versions == 1:
            cell.fill = red_fill
        elif total_versions == 2:
            cell.fill = green_fill
        elif total_versions > 2:
            cell.fill = blue_fill


wb.save(EXCEL_FILE)
print(f"\n✅ HRL files copied to '{HRL_PARENT_FOLDER}' and Excel file updated successfully!")